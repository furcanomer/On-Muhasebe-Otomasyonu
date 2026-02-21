import openpyxl
import os

path = r"c:\Users\furca\Desktop\ön muhasebe projesi\kaynak_excel_dosyası\ASAT DÖŞEMEALTI GELİR GİDER30.08.2025.xlsx"

try:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    print(f"=== SAYFALAR: {wb.sheetnames} ===\n")
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n{'='*60}")
        print(f"SAYFA: {sheet_name}  (max_row={ws.max_row}, max_col={ws.max_column})")
        print(f"{'='*60}")
        
        # İlk 30 satırı göster
        for i, row in enumerate(ws.iter_rows(max_row=30, values_only=True), 1):
            # Boş satırları atla
            if any(cell is not None for cell in row):
                print(f"  Satır {i:3d}: {row}")
        
        if ws.max_row > 30:
            print(f"  ... ({ws.max_row - 30} satır daha var)")
            # Son 5 satırı da göster
            all_rows = list(ws.iter_rows(values_only=True))
            for i, row in enumerate(all_rows[-5:], ws.max_row - 4):
                if any(cell is not None for cell in row):
                    print(f"  Satır {i:3d}: {row}")
    
    wb.close()

except Exception as e:
    print(f"Hata: {e}")
    import traceback
    traceback.print_exc()

input("\nDevam için Enter...")
