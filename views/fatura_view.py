import customtkinter as ctk
from tkinter import ttk, messagebox
import time
from datetime import datetime
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    has_reportlab = True
except ImportError:
    has_reportlab = False

class FaturaView(ctk.CTkFrame):
    def __init__(self, master, db):
        super().__init__(master, corner_radius=10, fg_color="transparent")
        self.db = db
        
        top_frame = ctk.CTkFrame(self, fg_color="transparent")
        top_frame.pack(fill="x", padx=10, pady=10)
        
        lbl = ctk.CTkLabel(top_frame, text="📄 Merkezi Evrak Yönetimi", font=ctk.CTkFont(family="Segoe UI Emoji", size=24, weight="bold"))
        lbl.pack(side="left")
        
        # Filtreleme / Görünüm Türü
        self.cmb_filter = ctk.CTkComboBox(top_frame, values=["Tümü", "Sipariş", "İrsaliye", "Fatura"], command=self.tabloyu_guncelle_cmd)
        self.cmb_filter.set("Tümü")
        self.cmb_filter.pack(side="left", padx=20)

        btn_yeni = ctk.CTkButton(top_frame, text="✚ Yeni Evrak Oluştur", command=self.evrak_ekle_modal, fg_color="#2ecc71", hover_color="#27ae60")
        btn_yeni.pack(side="right", padx=5)

        btn_donustur = ctk.CTkButton(top_frame, text="🔄 Seçiliyi Dönüştür", command=self.evrak_donustur_aksiyon, fg_color="orange", hover_color="darkorange")
        btn_donustur.pack(side="right", padx=5)

        btn_export = ctk.CTkButton(top_frame, text="📥 Seçiliyi Aktar", command=self.evrak_export_aksiyon, fg_color="#3498db", hover_color="#2980b9")
        btn_export.pack(side="right", padx=5)

        btn_export_all = ctk.CTkButton(top_frame, text="📂 Tümünü Aktar", command=self.tumunu_export_aksiyon, fg_color="#1abc9c", hover_color="#16a085")
        btn_export_all.pack(side="right", padx=5)

        btn_duzenle = ctk.CTkButton(top_frame, text="✏️ Düzenle", command=self.evrak_duzenle_aksiyon, fg_color="gray", hover_color="darkgray")
        btn_duzenle.pack(side="right", padx=5)

        btn_sil = ctk.CTkButton(top_frame, text="🗑️ Sil", command=self.fatura_sil_aksiyon, fg_color="red", hover_color="darkred")
        btn_sil.pack(side="right", padx=5)

        # Filtre Paneli (Tarih Aralığı)
        filter_frame = ctk.CTkFrame(self, fg_color="transparent")
        filter_frame.pack(fill="x", padx=15, pady=(0, 10))
        
        ctk.CTkLabel(filter_frame, text="Tarih Aralığı (YYYY-AA-GG):", font=ctk.CTkFont(size=12)).pack(side="left", padx=5)
        self.e_baslangic = ctk.CTkEntry(filter_frame, width=110, placeholder_text="0001-01-01")
        self.e_baslangic.pack(side="left", padx=5)
        self.e_baslangic.bind("<KeyRelease>", self._mask_date_event)
        
        ctk.CTkLabel(filter_frame, text="-").pack(side="left")
        
        self.e_bitis = ctk.CTkEntry(filter_frame, width=110, placeholder_text="2099-12-31")
        self.e_bitis.pack(side="left", padx=5)
        self.e_bitis.bind("<KeyRelease>", self._mask_date_event)
        
        btn_filtrele = ctk.CTkButton(filter_frame, text="🔍 Uygula", width=80, command=self.tabloyu_guncelle, fg_color="#34495e")
        btn_filtrele.pack(side="left", padx=10)
        
        btn_temizle = ctk.CTkButton(filter_frame, text="🧹", width=35, command=self.filtre_temizle, fg_color="gray")
        btn_temizle.pack(side="left", padx=2)

        table_frame = ctk.CTkFrame(self)
        table_frame.pack(fill="both", expand=True, padx=10, pady=10)

        self.tree = ttk.Treeview(table_frame, columns=("id", "fno", "cari", "tarih", "tutar", "yon", "tur", "odeme", "aciklama"), show="headings")
        self.tree.heading("id", text="ID")
        self.tree.heading("fno", text="Evrak No")
        self.tree.heading("cari", text="Cari")
        self.tree.heading("tarih", text="Tarih")
        self.tree.heading("tutar", text="Tutar (TL)")
        self.tree.heading("yon", text="Yön")
        self.tree.heading("tur", text="Tür")
        self.tree.heading("odeme", text="Ödeme")
        self.tree.heading("aciklama", text="Açıklama / Plaka")

        self.tree.column("id", width=40, anchor="center")
        self.tree.column("tutar", width=90, anchor="e")
        self.tree.column("yon", width=70, anchor="center")
        self.tree.column("tur", width=80, anchor="center")
        self.tree.pack(fill="both", expand=True, padx=5, pady=5)
        
        self.tabloyu_guncelle()

    def _mask_date_event(self, event):
        from utils import apply_date_mask
        apply_date_mask(event)

    def tabloyu_guncelle_cmd(self, val):
        self.tabloyu_guncelle()

    def tumunu_export_aksiyon(self):
        # Tablodaki tüm görünen satırları al
        items = self.tree.get_children()
        if not items:
            messagebox.showwarning("Uyarı", "Tabloda aktarılacak belge bulunamadı.")
            return
        
        id_list = [self.tree.item(i)['values'][0] for i in items]
        self.show_export_menu(id_list)

    def evrak_export_aksiyon(self):
        sel = self.tree.selection()
        if not sel: 
            messagebox.showwarning("Uyarı", "Lütfen dışa aktarmak istediğiniz evrağı/evrakları seçin.")
            return
        
        id_list = [self.tree.item(i)['values'][0] for i in sel]
        self.show_export_menu(id_list)

    def show_export_menu(self, id_list):
        # Export Menüsü
        m = ctk.CTkToplevel(self)
        m.title("Dışa Aktar")
        m.geometry("350x250")
        m.grab_set()
        m.attributes("-topmost", True)

        ctk.CTkLabel(m, text=f"{len(id_list)} Belge Seçildi", font=("Arial", 14)).pack(pady=10)
        ctk.CTkLabel(m, text="Format Seçin", font=("Arial", 16, "bold")).pack(pady=10)
        
        btn_pdf = ctk.CTkButton(m, text="📄 PDF (A4 Baskı) Olarak Kaydet", command=lambda: [m.destroy(), self.process_bulk_export(id_list, "pdf")])
        btn_pdf.pack(pady=5, padx=20, fill="x")
        
        btn_xl = ctk.CTkButton(m, text="📊 Excel (Tablo) Olarak Kaydet", command=lambda: [m.destroy(), self.process_bulk_export(id_list, "excel")])
        btn_xl.pack(pady=5, padx=20, fill="x")

    def process_bulk_export(self, id_list, format_type):
        success_count = 0
        total = len(id_list)
        
        last_err = ""
        for e_id in id_list:
            try:
                data = self.db.evrak_full_getir(e_id)
                if not data: 
                    last_err = f"Belge (ID: {e_id}) veritabanında bulunamadı."
                    continue
                
                if format_type == "pdf":
                    self.export_to_pdf(data, silent=True)
                else:
                    self.export_to_excel(data, silent=True)
                success_count += 1
            except Exception as e:
                last_err = str(e)
                print(f"Export Hatası (ID {e_id}): {e}")

        if success_count == total:
            messagebox.showinfo("Başarılı", f"{total} belgenin tamamı masaüstüne yüklendi.")
        elif success_count == 0:
            messagebox.showerror("Hata", f"Dışa aktarma başarısız oldu!\nHata Detayı: {last_err}")
        else:
            messagebox.showwarning("Kısmi Başarı", f"{total} belgeden {success_count} tanesi aktarıldı.\nSon Hata: {last_err}")

    def tr_fix(self, text):
        """PDF için Türkçe karakter düzeltme (Eğer font yüklenemezse)"""
        m = {"ı":"i", "İ":"I", "ğ":"g", "Ğ":"G", "ü":"u", "Ü":"U", "ş":"s", "Ş":"S", "ö":"o", "Ö":"O", "ç":"c", "Ç":"C"}
        for k, v in m.items():
            text = text.replace(k, v)
        return text

    def export_to_pdf(self, data, silent=False):
        if not has_reportlab:
            if not silent: messagebox.showerror("Hata", "PDF kütüphanesi eksik.")
            return

        h = data['header']
        det = data['detaylar']
        
        desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
        if not os.path.exists(desktop):
            # Masaüstü klasörü bulunamazsa (OneDrive vb.) ana kullanıcı klasörünü dene
            desktop = os.environ['USERPROFILE']

        # Dosya isminde geçersiz karakterleri temizle
        safe_fno = "".join([c for c in str(h[1]) if c.isalnum() or c in ('-', '_')])
        # h[6] Belge Türü (Fatura/İrsaliye vb) h[5] Yön (Alış/Satış)
        filename = f"{h[6]}_{safe_fno}.pdf"
        path = os.path.join(desktop, filename)
        
        try:
            # Türkçe Font Ayarı
            font_name = "Helvetica"
            try:
                # Arial varsa kullanalım
                arial_path = r"C:\Windows\Fonts\arial.ttf"
                if os.path.exists(arial_path):
                    pdfmetrics.registerFont(TTFont('ArialTurk', arial_path))
                    font_name = "ArialTurk"
            except: pass

            c = canvas.Canvas(path, pagesize=A4)
            width, height = A4
            
            def draw_t(canvas, x, y, text, size=10, bold=False):
                f = font_name + ("-Bold" if bold and font_name=="Helvetica" else "")
                canvas.setFont(f, size)
                # Arial değilse karakterleri temizle
                if font_name == "Helvetica":
                    text = self.tr_fix(str(text))
                canvas.drawString(x, y, str(text))

            # header indices: 0:id, 1:f_no... 10:aciklama, 11:unvan, 12:vn, 13:vd, 14:adr
            # detaylar indices: 0:id, 1:fid, 2:sid, 3:mik, 4:price, 5:kdv_r, 6:tvk_r, 7:birim, 8:name, 9:barcode
            
            # Başlık
            draw_t(c, 50, height - 50, f"{h[6]} Belgesi", 18, True)
            
            draw_t(c, 50, height - 80, f"Belge No: {h[1]}")
            draw_t(c, 50, height - 95, f"Tarih: {h[3]}")
            draw_t(c, 50, height - 110, f"Tür: {h[5]} / Ödeme: {h[9]}")
            draw_t(c, 50, height - 125, f"Plaka / Açıklama: {h[10] or '-'}")
            
            # Cari Bölümü
            c.rect(300, height - 130, 250, 70)
            draw_t(c, 310, height - 75, "SAYIN (CARİ):", 10, True)
            draw_t(c, 310, height - 95, f"{h[11]}", 11, True) # unvan
            draw_t(c, 310, height - 110, f"V.N: {h[12]} / V.D: {h[13]}") # vergi_no / vergi_dairesi
            draw_t(c, 310, height - 122, f"Adres: {h[14]}") # adres
            
            # Tablo
            y = height - 160
            c.line(50, y, 550, y)
            draw_t(c, 55, y - 15, "Ürün / Hizmet", 9, True)
            draw_t(c, 240, y - 15, "Miktar", 9, True)
            draw_t(c, 300, y - 15, "Birim", 9, True)
            draw_t(c, 360, y - 15, "B.Fiyat", 9, True)
            draw_t(c, 420, y - 15, "KDV/Tevk", 9, True)
            draw_t(c, 500, y - 15, "Net", 9, True)
            y -= 25
            c.line(50, y, 550, y)
            
            y -= 15
            for d in det:
                # 3:mik, 4:price, 5:kdv, 6:tvk, 7:unit, 8:name
                kdvsiz = d[3]*d[4]
                kdv_t = kdvsiz * d[5] / 100
                tvk_t = kdvsiz * d[6] / 100
                net = kdvsiz + kdv_t - tvk_t
                
                draw_t(c, 55, y, d[8]) # name
                draw_t(c, 240, y, f"{d[3]}") # mik
                draw_t(c, 300, y, f"{d[7]}") # unit
                draw_t(c, 360, y, f"{d[4]:.2f}") # price
                draw_t(c, 420, y, f"%{d[5]:.0f}/%{d[6]:.0f}") # kdv/tevk
                draw_t(c, 500, y, f"{net:.2f}") # net
                y -= 18
                if y < 100:
                    c.showPage()
                    y = height - 50
            
            c.line(50, y, 550, y)
            draw_t(c, 400, y - 30, f"GENEL TOPLAM:", 12, True)
            draw_t(c, 500, y - 30, f"{h[4]:.2f} TL", 12, True)
            
            c.save()
            if not silent: messagebox.showinfo("Başarılı", f"PDF Kaydedildi: {filename}")
        except Exception as e:
            if not silent: messagebox.showerror("Hata", f"PDF oluşturulamadı: {e}")
            raise e

    def export_to_excel(self, data, silent=False):
        h = data['header']
        det = data['detaylar']
        
        desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
        if not os.path.exists(desktop): desktop = os.environ['USERPROFILE']
        
        safe_fno = "".join([c for c in str(h[1]) if c.isalnum() or c in ('-', '_')])
        filename = f"{h[6]}_{safe_fno}.xlsx"
        path = os.path.join(desktop, filename)
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = str(h[5])
            
            # Başlıklar
            ws.merge_cells('A1:E1')
            ws['A1'] = f"{h[5]} DETAYI"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # header indices: 0:id, 1:f_no, 2:cid, 3:date, 4:total, 5:tur, 6:b_turu, 10:aciklama, 11:unvan
            # detaylar indices: 0:id, 1:f_id, 2:s_id, 3:mik, 4:price, 5:kdv, 6:tevk, 7:birim, 8:name, 9:barcode
            
            ws.append(["Evrak No", h[1], "", "Tarih", h[3]])
            ws.append(["Cari", h[11], "", "Plaka/Açkl", h[10]]) # unvan ve aciklama
            ws.append([])
            
            ws.append(["Ürün Adı", "Miktar", "Birim", "B.Fiyat", "KDVsiz", "KDV %", "Tevk %", "Net Toplam"])
            kdvsiz_toplam = 0
            kdv_toplam = 0
            tevk_toplam = 0
            
            for d in det:
                mik = d[3]; bf = d[4]; kdv_r = d[5]; tevk_r = d[6]; unit = d[7]; name = d[8]
                line_kdvsiz = mik * bf
                line_kdv = line_kdvsiz * kdv_r / 100
                line_tevk = line_kdvsiz * tevk_r / 100
                line_net = line_kdvsiz + line_kdv - line_tevk
                
                ws.append([name, mik, unit, bf, line_kdvsiz, f"%{kdv_r}", f"%{tevk_r}", line_net])
                kdvsiz_toplam += line_kdvsiz
                kdv_toplam += line_kdv
                tevk_toplam += line_tevk
                
            ws.append([])
            ws.append(["", "", "", "", "", "", "KDVsiz Toplam", kdvsiz_toplam])
            ws.append(["", "", "", "", "", "", "KDV (+)", kdv_toplam])
            ws.append(["", "", "", "", "", "", "Tevkifat (-)", tevk_toplam])
            ws.append(["", "", "", "", "", "", "GENEL TOPLAM", h[4]])
            
            # Stil
            for r in range(ws.max_row - 3, ws.max_row + 1):
                ws.cell(row=r, column=7).font = Font(bold=True)
                ws.cell(row=r, column=8).font = Font(bold=True)
            
            # Sütun Genişliği
            for col in ['A', 'D', 'H']:
                ws.column_dimensions[col].width = 18

            wb.save(path)
            if not silent: messagebox.showinfo("Başarılı", f"Excel Kaydedildi: {filename}")
        except Exception as e:
            if not silent: messagebox.showerror("Hata", f"Excel hatası: {e}")
            raise e

    def tabloyu_guncelle(self):
        for row in self.tree.get_children(): self.tree.delete(row)
        filtre = self.cmb_filter.get()
        if filtre == "Tümü": filtre = None
        
        bas = self.e_baslangic.get().strip() or None
        bit = self.e_bitis.get().strip() or None
        
        for f in self.db.tum_faturalari_getir(belge_turu=filtre, baslangic=bas, bitis=bit): 
            self.tree.insert("", "end", values=f)

    def filtre_temizle(self):
        self.e_baslangic.delete(0, "end")
        self.e_bitis.delete(0, "end")
        self.cmb_filter.set("Tümü")
        self.tabloyu_guncelle()

    def evrak_donustur_aksiyon(self):
        sel = self.tree.selection()
        if not sel: return
        item = self.tree.item(sel[0])['values']
        e_id, e_turu = item[0], item[6]
        
        if e_turu == "Fatura":
            messagebox.showinfo("Bilgi", "Fatura en üst seviye belgedir, daha ileri dönüştürülemez.")
            return

        hedef = "İrsaliye" if e_turu == "Sipariş" else "Fatura"
        if messagebox.askyesno("Dönüştür", f"Seçili {e_turu} belgesini {hedef} belgesine dönüştürmek istiyor musunuz?"):
            with self.master.master.islem_bekle("Dönüştürülüyor..."):
                if self.db.evrak_donustur(e_id, hedef):
                    messagebox.showinfo("Başarılı", f"Belge {hedef} olarak oluşturuldu.")
                    self.tabloyu_guncelle()
                else:
                    messagebox.showerror("Hata", "Dönüştürme başarısız.")

    def evrak_duzenle_aksiyon(self):
        selected = self.tree.selection()
        if not selected: return
        self.evrak_ekle_modal(edit_id=self.tree.item(selected[0])['values'][0])

    def fatura_sil_aksiyon(self):
        selected = self.tree.selection()
        if not selected: return
        if messagebox.askyesno("Onay", "Silmek istediğinize emin misiniz?"):
            if self.db.fatura_sil(self.tree.item(selected[0])['values'][0]):
                self.tabloyu_guncelle()

    def evrak_ekle_modal(self, edit_id=None):
        cariler = self.db.tum_carileri_getir()
        stoklar = self.db.tum_stoklari_getir()
        if not cariler or not stoklar:
            messagebox.showerror("Hata", "Önce Cari ve Stok tanımlamalısınız.")
            return

        c_dict = {f"{c[1]}": c for c in cariler}
        s_barcode_dict = {f"{s[1]}": s for s in stoklar} # barkod -> stok
        s_name_dict = {f"{s[2]}": s for s in stoklar}

        edit_data = None
        edit_detaylar = []
        if edit_id:
            with self.db._connect() as conn:
                edit_data = conn.execute("SELECT fatura_no, cari_id, tarih, tur, belge_turu, fiyat_goster, odeme_turu, aciklama FROM faturalar WHERE id=?", (edit_id,)).fetchone()
                edit_detaylar = conn.execute("SELECT stok_id, miktar, birim_fiyat, kdv_orani, tevkifat_orani, birim FROM fatura_detay WHERE fatura_id=?", (edit_id,)).fetchall()

        p = ctk.CTkToplevel(self)
        p.title("Evrak Oluştur/Düzenle")
        p.geometry("800x750")
        p.grab_set()

        # Üst Panel: Belge Ayarları
        conf_frame = ctk.CTkFrame(p)
        conf_frame.pack(fill="x", padx=10, pady=10)

        ctk.CTkLabel(conf_frame, text="Evrak No:").grid(row=0, column=0, padx=5, pady=5)
        e_fno = ctk.CTkEntry(conf_frame, width=120)
        e_fno.insert(0, edit_data[0] if edit_id else f"EVR-{int(time.time())}")
        e_fno.grid(row=0, column=1, padx=5, pady=5)

        ctk.CTkLabel(conf_frame, text="Tür:").grid(row=0, column=2, padx=5, pady=5)
        cmb_tur = ctk.CTkComboBox(conf_frame, values=["Sipariş", "İrsaliye", "Fatura"], width=100)
        cmb_tur.set(edit_data[4] if edit_id else "Fatura")
        cmb_tur.grid(row=0, column=3, padx=5, pady=5)

        ctk.CTkLabel(conf_frame, text="Yön:").grid(row=0, column=4, padx=5, pady=5)
        cmb_yon = ctk.CTkComboBox(conf_frame, values=["Satış", "Alış"], width=80)
        cmb_yon.set(edit_data[3] if edit_id else "Satış")
        cmb_yon.grid(row=0, column=5, padx=5, pady=5)

        ctk.CTkLabel(conf_frame, text="Ödeme:").grid(row=0, column=6, padx=5, pady=5)
        cmb_odem = ctk.CTkComboBox(conf_frame, values=["Veresiye", "Nakit"], width=90)
        cmb_odem.set(edit_data[6] if edit_id else "Veresiye")
        cmb_odem.grid(row=0, column=7, padx=5, pady=5)

        ctk.CTkLabel(conf_frame, text="Plaka/Açkl:").grid(row=1, column=0, padx=5, pady=5)
        e_aciklama = ctk.CTkEntry(conf_frame, width=120)
        e_aciklama.insert(0, edit_data[7] if edit_id and edit_data[7] else "")
        e_aciklama.grid(row=1, column=1, padx=5, pady=5)

        ctk.CTkLabel(conf_frame, text="Proje Kodu:").grid(row=1, column=2, padx=5, pady=5)
        e_proje = ctk.CTkEntry(conf_frame, width=100)
        # edit_data indexleri: 0:fno, 1:cari_id, 2:tarih, 3:tur, 4:belge_turu, 5:fiyat_goster, 6:odeme_turu, 7:aciklama, 8:proje, 9:yuvarlama, 10:tevk_turu
        # NOT: edit_data'yı güncelleyelim ilerde. Şimdilik manuel.
        with self.db._connect() as conn:
            extra = conn.execute("SELECT proje_kodu, yuvarlama_farki, tevkifat_turu FROM faturalar WHERE id=?", (edit_id or 0,)).fetchone()
        
        e_proje.insert(0, extra[0] if extra and extra[0] else "GENEL")
        e_proje.grid(row=1, column=3, padx=5, pady=5)

        ctk.CTkLabel(conf_frame, text="Yuv. Farkı:").grid(row=1, column=4, padx=5, pady=5)
        e_yuvarlama = ctk.CTkEntry(conf_frame, width=60)
        e_yuvarlama.insert(0, str(extra[1]) if extra else "0.0")
        e_yuvarlama.grid(row=1, column=5, padx=5, pady=5)

        ctk.CTkLabel(conf_frame, text="Tevk. Tür:").grid(row=1, column=6, padx=5, pady=5)
        cmb_tevk_turu = ctk.CTkComboBox(conf_frame, values=["", "2/10", "3/10", "4/10", "5/10", "7/10", "9/10"], width=80)
        cmb_tevk_turu.set(extra[2] if extra and extra[2] else "")
        cmb_tevk_turu.grid(row=1, column=7, padx=5, pady=5)

        check_fiyat = ctk.CTkCheckBox(conf_frame, text="Fiyatları Göster")
        if edit_id:
            if edit_data[5]: check_fiyat.select()
            else: check_fiyat.deselect()
        else: check_fiyat.select()
        check_fiyat.grid(row=1, column=0, columnspan=2, padx=5, pady=5)

        ctk.CTkLabel(conf_frame, text="Cari Seçimi:").grid(row=1, column=2, padx=5, pady=5)
        combo_cari = ctk.CTkComboBox(conf_frame, values=list(c_dict.keys()), width=200)
        if edit_id:
            for k, v in c_dict.items():
                if v[0] == edit_data[1]: combo_cari.set(k); break
        combo_cari.grid(row=1, column=3, columnspan=2, padx=5, pady=5)

        # Ürün Ekleme Satırı
        entry_frame = ctk.CTkFrame(p)
        entry_frame.pack(fill="x", padx=10, pady=5)

        ctk.CTkLabel(entry_frame, text="Ürün:").pack(side="left", padx=5)
        combo_urun = ctk.CTkComboBox(entry_frame, values=list(s_name_dict.keys()), width=200)
        combo_urun.pack(side="left", padx=5, pady=10)

        ctk.CTkLabel(entry_frame, text="Miktar:").pack(side="left", padx=5)
        e_mik = ctk.CTkEntry(entry_frame, width=60)
        e_mik.insert(0, "1")
        e_mik.pack(side="left", padx=5)

        ctk.CTkLabel(entry_frame, text="KDV %:").pack(side="left", padx=5)
        cmb_kdv = ctk.CTkComboBox(entry_frame, values=["0", "1", "10", "20"], width=55)
        cmb_kdv.set("20")
        cmb_kdv.pack(side="left", padx=5)

        ctk.CTkLabel(entry_frame, text="Tevkifat %:").pack(side="left", padx=5)
        cmb_tevkifat = ctk.CTkComboBox(entry_frame, values=["0", "20", "30", "40", "50", "60", "70", "80", "90"], width=55)
        cmb_tevkifat.set("0")
        cmb_tevkifat.pack(side="left", padx=5)

        def kalem_ekle_aksiyon():
            try:
                name = combo_urun.get()
                if name not in s_name_dict:
                    messagebox.showerror("Hata", "Lütfen listeden bir ürün seçin.")
                    return
                stok = s_name_dict[name]
                qty = float(e_mik.get().replace(",","."))
                if qty <= 0: raise ValueError
                kdv_r = float(cmb_kdv.get())
                tvk_r = float(cmb_tevkifat.get())
                add_item_to_list(stok, qty, kdv_r, tvk_r)
            except Exception as e:
                messagebox.showerror("Hata", "Geçerli bir miktar giriniz.")

        btn_ekle_kalem = ctk.CTkButton(entry_frame, text="✚ Listeye Ekle", width=100, command=kalem_ekle_aksiyon)
        btn_ekle_kalem.pack(side="left", padx=10)

        # Kalemler Tablosu
        list_frame = ctk.CTkFrame(p)
        list_frame.pack(fill="both", expand=True, padx=10, pady=5)
        
        cols = ("stok_id", "barkod", "urun", "miktar", "birim_fiyat", "kdv", "tevkifat", "kdvsiz", "kdvli", "net")
        tree_k = ttk.Treeview(list_frame, columns=cols, show="headings")
        tree_k.heading("barkod", text="Barkod")
        tree_k.heading("urun", text="Ürün Adı")
        tree_k.heading("miktar", text="Miktar")
        tree_k.heading("birim_fiyat", text="B.Fiyat")
        tree_k.heading("kdv", text="KDV %")
        tree_k.heading("tevkifat", text="Tevkifat %")
        tree_k.heading("kdvsiz", text="KDVsiz")
        tree_k.heading("kdvli", text="KDV Tutarı")
        tree_k.heading("net", text="Net Tutar")
        tree_k.column("stok_id", width=0, stretch=False)
        tree_k.column("barkod", width=90)
        tree_k.column("urun", width=180)
        tree_k.column("miktar", width=55, anchor="center")
        tree_k.column("birim_fiyat", width=65, anchor="e")
        tree_k.column("kdv", width=45, anchor="center")
        tree_k.column("tevkifat", width=65, anchor="center")
        tree_k.column("kdvsiz", width=80, anchor="e")
        tree_k.column("kdvli", width=80, anchor="e")
        tree_k.column("net", width=80, anchor="e")
        tree_k.pack(fill="both", expand=True)

        totals_frame = ctk.CTkFrame(p, fg_color="transparent")
        totals_frame.pack(fill="x", padx=20, pady=5)
        lbl_kdvsiz_t = ctk.CTkLabel(totals_frame, text="KDVsiz: 0.00", font=("Arial", 11))
        lbl_kdvsiz_t.pack(side="left", padx=10)
        lbl_kdv_t = ctk.CTkLabel(totals_frame, text="KDV: 0.00", font=("Arial", 11), text_color="orange")
        lbl_kdv_t.pack(side="left", padx=10)
        lbl_tevkifat_t = ctk.CTkLabel(totals_frame, text="Tevkifat: 0.00", font=("Arial", 11), text_color="#e74c3c")
        lbl_tevkifat_t.pack(side="left", padx=10)
        lbl_total = ctk.CTkLabel(totals_frame, text="TOPLAM: 0.00 TL", font=("Arial", 15, "bold"), text_color="#2ecc71")
        lbl_total.pack(side="right", padx=10)

        self.detaylar = [] # (stok_id, miktar, birim_fiyat, kdv_orani, tevkifat_orani, birim)

        def reflect_total():
            kdvsiz = sum([float(d[1])*float(d[2]) for d in self.detaylar])
            kdv_t = sum([float(d[1])*float(d[2])*float(d[3])/100 for d in self.detaylar])
            tevk_t = sum([float(d[1])*float(d[2])*float(d[4])/100 for d in self.detaylar])
            net = kdvsiz + kdv_t - tevk_t
            lbl_kdvsiz_t.configure(text=f"KDVsiz: {kdvsiz:.2f}")
            lbl_kdv_t.configure(text=f"KDV: +{kdv_t:.2f}")
            lbl_tevkifat_t.configure(text=f"Tevkifat: -{tevk_t:.2f}")
            yuv = float(e_yuvarlama.get() or 0)
            lbl_total.configure(text=f"TOPLAM: {net + yuv:.2f} TL")

        def add_item_to_list(stok, qty, kdv_r=20.0, tvk_r=0.0):
             cari_unvan = combo_cari.get()
             iskonto = 0.0
             if cari_unvan in c_dict:
                 iskonto = c_dict[cari_unvan][7] # ozel_iskonto
             
             base_price = stok[5]
             final_price = base_price * (1 - (iskonto / 100))
             birim_str = stok[6] if len(stok) > 6 else 'Adet'
             
             kdvsiz = qty * final_price
             kdv_tutari = kdvsiz * kdv_r / 100
             tevk_tutari = kdvsiz * tvk_r / 100
             net = kdvsiz + kdv_tutari - tevk_tutari
             
             self.detaylar.append([stok[0], qty, final_price, kdv_r, tvk_r, birim_str])
             tree_k.insert("", "end", values=(stok[0], stok[1], stok[2], qty, f"{final_price:.2f}", f"%{kdv_r:.0f}", f"%{tvk_r:.0f}", f"{kdvsiz:.2f}", f"{kdv_tutari:.2f}", f"{net:.2f}"))
             reflect_total()

        if edit_id:
            for row in edit_detaylar:
                sid = row[0]; mik = row[1]; bfiy = row[2]
                kdv_r = row[3] if len(row) > 3 else 0.0
                tvk_r = row[4] if len(row) > 4 else 0.0
                birim_str = row[5] if len(row) > 5 else 'Adet'
                s_obj = None
                for s in stoklar:
                    if s[0] == sid: s_obj = s; break
                if s_obj:
                    kdvsiz = mik * bfiy
                    kdv_t = kdvsiz * kdv_r / 100
                    tvk_t = kdvsiz * tvk_r / 100
                    self.detaylar.append([sid, mik, bfiy, kdv_r, tvk_r, birim_str])
                    tree_k.insert("", "end", values=(sid, s_obj[1], s_obj[2], mik, bfiy, f"%{kdv_r:.0f}", f"%{tvk_r:.0f}", f"{kdvsiz:.2f}", f"{kdv_t:.2f}", f"{kdvsiz+kdv_t-tvk_t:.2f}"))
            reflect_total()

        def save_all():
            if not self.detaylar:
                messagebox.showwarning("Uyarı", "Kalem listesi boş. En az bir ürün ekleyin.")
                return
            if combo_cari.get() not in c_dict:
                messagebox.showwarning("Uyarı", "Lütfen geçerli bir cari seçin.")
                return
            try:
                cid = c_dict[combo_cari.get()][0]
                fno = e_fno.get()
                bturu = cmb_tur.get()
                yon = cmb_yon.get()
                odem = cmb_odem.get()
                acik = e_aciklama.get()
                f_goster = 1 if check_fiyat.get() else 0
                
                with self.master.master.islem_bekle():
                    if edit_id:
                        res = self.db.fatura_guncelle(edit_id, fno, cid, yon, self.detaylar, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), belge_turu=bturu, fiyat_goster=f_goster, odeme_turu=odem, aciklama=acik, proje_kodu=e_proje.get(), yuvarlama_farki=float(e_yuvarlama.get() or 0), tevkifat_turu=cmb_tevk_turu.get())
                    else:
                        res = self.db.fatura_olustur(fno, cid, yon, self.detaylar, belge_turu=bturu, fiyat_goster=f_goster, odeme_turu=odem, aciklama=acik, proje_kodu=e_proje.get(), yuvarlama_farki=float(e_yuvarlama.get() or 0), tevkifat_turu=cmb_tevk_turu.get())
                    
                    if res:
                        messagebox.showinfo("Başarılı", "Evrak kaydedildi.")
                        self.tabloyu_guncelle()
                        p.destroy()
            except Exception as e:
                messagebox.showerror("Hata", str(e))

        btn_save = ctk.CTkButton(p, text="💾 EVRAĞI KAYDET (F2)", height=45, command=save_all)
        btn_save.pack(pady=20)
        p.bind("<F2>", lambda e: save_all())
        p.bind("<Escape>", lambda e: p.destroy())
        ctk.CTkButton(p, text="❌ Satır Sil", fg_color="red", command=lambda: [tree_k.delete(s) for s in tree_k.selection()]).pack(pady=5)
