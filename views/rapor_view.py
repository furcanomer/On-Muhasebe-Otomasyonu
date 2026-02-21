import customtkinter as ctk
from tkinter import ttk, messagebox
import os
from datetime import datetime
try:
    import openpyxl
    from openpyxl.styles import Font
except ImportError:
    openpyxl = None

class RaporView(ctk.CTkFrame):
    def __init__(self, master, db):
        super().__init__(master, corner_radius=10, fg_color="transparent")
        self.db = db
        
        # Üst Panel
        top_frame = ctk.CTkFrame(self, fg_color="transparent")
        top_frame.pack(fill="x", padx=10, pady=10)
        
        lbl = ctk.CTkLabel(top_frame, text="📊 Genel Raporlama", font=ctk.CTkFont(family="Segoe UI Emoji", size=24, weight="bold"))
        lbl.pack(side="left")

        btn_yenile = ctk.CTkButton(top_frame, text="🔄 Verileri Güncelle", command=self.gosterge_paneli_guncelle, fg_color="gray", hover_color="darkgray")
        btn_yenile.pack(side="right", padx=5)

        # GÜNLÜK ÖZET PANELİ (Yeni İstek)
        gunluk_frame = ctk.CTkFrame(self, corner_radius=10, border_width=2, border_color="#3498db")
        gunluk_frame.pack(fill="x", padx=20, pady=10)
        
        ctk.CTkLabel(gunluk_frame, text="📅 GÜNLÜK OPERASYON ÖZETİ (BUGÜN)", font=ctk.CTkFont(size=16, weight="bold")).pack(pady=5)
        
        inner_gunluk = ctk.CTkFrame(gunluk_frame, fg_color="transparent")
        inner_gunluk.pack(fill="x", padx=10, pady=10)
        
        self.lbl_gunluk_ciro = self.kart_yap(inner_gunluk, "Günlük Toplam Ciro", "0.00 TL", "#2ecc71", 0)
        self.lbl_gunluk_nakit = self.kart_yap(inner_gunluk, "Günlük Nakit Tahsilat", "#3498db", "blue", 1)
        self.lbl_gunluk_veresiye = self.kart_yap(inner_gunluk, "Günlük Veresiye Satış", "#e67e22", "orange", 2)
        inner_gunluk.columnconfigure((0, 1, 2), weight=1)

        # GENEL DURUM PANELİ
        ctk.CTkLabel(self, text="🏛 GENEL FİNANSAL DURUM", font=ctk.CTkFont(size=16, weight="bold")).pack(pady=(10, 0))
        self.kart_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.kart_frame.pack(fill="x", padx=20, pady=10)
        
        self.lbl_alacak = self.kart_yap(self.kart_frame, "Toplam Alacaklar", "0.00 TL", "green", 0)
        self.lbl_borc = self.kart_yap(self.kart_frame, "Toplam Borçlar", "0.00 TL", "red", 1)
        self.lbl_kasa = self.kart_yap(self.kart_frame, "Toplam Kasa & Banka", "0.00 TL", "blue", 2)
        self.lbl_stok = self.kart_yap(self.kart_frame, "Stok Değeri", "0.00 TL", "orange", 3)
        self.kart_frame.columnconfigure((0, 1, 2, 3), weight=1)

        # KDV / TEVKİFAT ÖZET PANELİ
        kdv_frame = ctk.CTkFrame(self, corner_radius=10, border_width=2, border_color="#e67e22")
        kdv_frame.pack(fill="x", padx=20, pady=5)

        ctk.CTkLabel(kdv_frame, text="🧾 KDV / TEVKİFAT ÖZETİ", font=ctk.CTkFont(size=14, weight="bold")).pack(pady=5)
        inner_kdv = ctk.CTkFrame(kdv_frame, fg_color="transparent")
        inner_kdv.pack(fill="x", padx=10, pady=5)
        self.lbl_kdvsiz = self.kart_yap(inner_kdv, "KDVsiz Matrah", "0.00 TL", "#3498db", 0)
        self.lbl_kdv_hesap = self.kart_yap(inner_kdv, "Hesaplanan KDV", "0.00 TL", "#e67e22", 1)
        self.lbl_tevkifat = self.kart_yap(inner_kdv, "Tevkifat Tutarı", "0.00 TL", "#e74c3c", 2)
        self.lbl_odenecek_kdv = self.kart_yap(inner_kdv, "Ödenecek Net KDV", "0.00 TL", "#2ecc71", 3)
        inner_kdv.columnconfigure((0, 1, 2, 3), weight=1)

        # Dışa Aktar Çerçevesi
        export_frame = ctk.CTkFrame(self, corner_radius=10)
        export_frame.pack(fill="x", padx=20, pady=20)

        ctk.CTkLabel(export_frame, text="📥 VERİLERİ DIŞA AKTAR (GERÇEK EXCEL FORMATI)", font=ctk.CTkFont(size=14, weight="bold")).pack(pady=10)
        
        btn_frame = ctk.CTkFrame(export_frame, fg_color="transparent")
        btn_frame.pack(pady=10)
        
        ctk.CTkButton(btn_frame, text="👥 Cari Listesi", command=self.export_cariler).pack(side="left", padx=5)
        ctk.CTkButton(btn_frame, text="📦 Stok Listesi", command=self.export_stoklar).pack(side="left", padx=5)
        ctk.CTkButton(btn_frame, text="💵 Kasa Hareketleri", command=self.export_kasa_h).pack(side="left", padx=5)
        ctk.CTkButton(btn_frame, text="📑 Evrak Listesi", command=self.export_evraklar).pack(side="left", padx=5)

        self.gosterge_paneli_guncelle()


    def kart_yap(self, parent, baslik, deger, renk, col):
        frame = ctk.CTkFrame(parent, corner_radius=15, border_width=1)
        frame.grid(row=0, column=col, padx=10, pady=10, sticky="nsew")
        
        ctk.CTkLabel(frame, text=baslik, font=ctk.CTkFont(size=13), text_color="gray").pack(pady=(10, 2))
        lbl_deger = ctk.CTkLabel(frame, text=deger, font=ctk.CTkFont(size=18, weight="bold"), text_color=renk)
        lbl_deger.pack(pady=(0, 15))
        return lbl_deger

    def gosterge_paneli_guncelle(self):
        # Genel Durum
        durum = self.db.genel_durum_getir()
        self.lbl_alacak.configure(text=f"{durum['Alacaklar']:.2f} TL")
        self.lbl_borc.configure(text=f"{durum['Bor\u00e7lar']:.2f} TL")
        self.lbl_kasa.configure(text=f"{durum['Kasa_Banka']:.2f} TL")
        self.lbl_stok.configure(text=f"{durum['Stok_Degeri']:.2f} TL")

        # Günlük Özet
        gunluk = self.db.gunluk_ozet_getir()
        self.lbl_gunluk_ciro.configure(text=f"{gunluk['ciro']:.2f} TL")
        self.lbl_gunluk_nakit.configure(text=f"{gunluk['nakit']:.2f} TL")
        self.lbl_gunluk_veresiye.configure(text=f"{gunluk['veresiye']:.2f} TL")

        # KDV Özeti
        kdv_rows = self.db.kdv_ozet_getir()
        kdvsiz_t = sum(r[0] or 0 for r in kdv_rows)
        kdv_t = sum(r[1] or 0 for r in kdv_rows)
        tevk_t = sum(r[2] or 0 for r in kdv_rows)
        odenecek = kdv_t - tevk_t
        self.lbl_kdvsiz.configure(text=f"{kdvsiz_t:.2f} TL")
        self.lbl_kdv_hesap.configure(text=f"{kdv_t:.2f} TL")
        self.lbl_tevkifat.configure(text=f"{tevk_t:.2f} TL")
        self.lbl_odenecek_kdv.configure(text=f"{odenecek:.2f} TL")

    def excel_kaydet(self, sql_sorgu, dosya_adi):
        if not openpyxl:
            messagebox.showerror("Hata", "openpyxl kütüphanesi yüklü değil!")
            return
            
        try:
            with self.db._connect() as conn:
                cursor = conn.cursor()
                cursor.execute(sql_sorgu)
                rows = cursor.fetchall()
                columns = [description[0].replace('_', ' ').title() for description in cursor.description]
                
                desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
                tam_yol = os.path.join(desktop, dosya_adi)
                
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Rapor"
                
                # Başlıklar
                ws.append(columns)
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                
                # Veriler
                for r in rows:
                    ws.append(r)
                
                # Sütun Genişlikleri (#### hatasını önlemek için)
                for idx, col in enumerate(ws.columns, 1):
                    ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = 20
                
                wb.save(tam_yol)
                messagebox.showinfo("Başarılı", f"Excel Masaüstüne Kaydedildi:\n{dosya_adi}")
        except Exception as e:
            messagebox.showerror("Hata", f"Dışa aktarım hatası:\n{e}")

    def export_cariler(self):
        self.excel_kaydet("SELECT unvan, vergi_no, telefon, bakiye, cari_turu, ozel_iskonto, fiyat_grubu FROM cariler", "Cari_Listesi_Rapor.xlsx")

    def export_stoklar(self):
        self.excel_kaydet("SELECT barkod, urun_adi, miktar, fiyat FROM stoklar", "Stok_Listesi_Rapor.xlsx")

    def export_kasa_h(self):
        self.excel_kaydet("""
            SELECT kh.tarih, kb.hesap_adi as kasa_banka, ifnull(c.unvan, 'Genel') as cari, kh.tutar, kh.tur, kh.aciklama 
            FROM kasa_hareketleri kh
            JOIN kasalar kb ON kh.kasa_id = kb.id
            LEFT JOIN cariler c ON kh.cari_id = c.id
            UNION ALL
            SELECT bh.tarih, b.hesap_adi as kasa_banka, ifnull(c.unvan, 'Genel') as cari, bh.tutar, bh.tur, bh.aciklama
            FROM banka_hareketleri bh
            JOIN bankalar b ON bh.banka_id = b.id
            LEFT JOIN cariler c ON bh.cari_id = c.id
            ORDER BY 1 DESC
        """, "Kasa_ve_Banka_Hareketleri_Rapor.xlsx")

    def export_evraklar(self):
        self.excel_kaydet("""
            SELECT f.tarih, f.belge_turu, f.fatura_no, c.unvan, f.tur as yon, f.toplam_tutar, f.odeme_turu, f.aciklama as plaka
            FROM faturalar f
            JOIN cariler c ON f.cari_id = c.id
            ORDER BY f.tarih DESC
        """, "Tum_Evraklar_Rapor.xlsx")
