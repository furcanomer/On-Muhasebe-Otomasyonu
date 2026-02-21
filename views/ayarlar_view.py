import customtkinter as ctk
from tkinter import messagebox, filedialog
from datetime import datetime
import os

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None

class AyarlarView(ctk.CTkFrame):
    def __init__(self, master, master_db, main_app):
        super().__init__(master, corner_radius=10, fg_color="transparent")
        self.master_db = master_db
        self.main_app = main_app
        
        # Başlık
        top_frame = ctk.CTkFrame(self, fg_color="transparent")
        top_frame.pack(fill="x", padx=10, pady=10)
        
        lbl = ctk.CTkLabel(top_frame, text="⚙️ Ayarlar", font=ctk.CTkFont(family="Segoe UI Emoji", size=24, weight="bold"))
        lbl.pack(side="left")

        # Ayar Kartı
        settings_frame = ctk.CTkFrame(self, corner_radius=15)
        settings_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Tema Ayarı
        ctk.CTkLabel(settings_frame, text="Görünüm Modu (Tema)", font=ctk.CTkFont(size=16, weight="bold")).pack(pady=(30, 10), padx=30, anchor="w")
        
        current_theme = self.master_db.ayar_getir("tema", "Dark")
        
        self.theme_option = ctk.CTkOptionMenu(
            settings_frame, 
            values=["Light", "Dark"],
            command=self.change_theme_event
        )
        self.theme_option.set(current_theme)
        self.theme_option.pack(pady=10, padx=30, anchor="w")
        
        ctk.CTkLabel(
            settings_frame, 
            text="Uygulama temasını değiştirdiğinizde tablolar ve renkler anlık olarak güncellenir.",
            font=ctk.CTkFont(size=12),
            text_color="gray"
        ).pack(pady=5, padx=30, anchor="w")

        # Bölücü çizgi
        ctk.CTkFrame(settings_frame, height=2, fg_color=("gray80", "gray30")).pack(fill="x", padx=30, pady=20)

        # Excel Dönüştürme
        excel_frame = ctk.CTkFrame(settings_frame, fg_color="transparent")
        excel_frame.pack(fill="x", padx=30, pady=(0, 10))
        
        ctk.CTkLabel(excel_frame, text="📊 Veri Dışa Aktarımı", font=ctk.CTkFont(size=16, weight="bold")).pack(anchor="w", pady=(0, 5))
        
        ctk.CTkLabel(
            excel_frame, 
            text="Soldaki menüdeki tüm bölümler (Cariler, Stoklar, Kasa, Banka, Evraklar, Personel, Takvim, Notlar)\nayrı sekmeler olarak tek bir Excel dosyasına aktarılır. Excel ile yola devam edebilirsiniz.",
            font=ctk.CTkFont(size=12),
            text_color="gray",
            justify="left"
        ).pack(anchor="w", pady=(0, 10))

        btn_excel_aktar = ctk.CTkButton(
            excel_frame, 
            text="�  Bu Şirketi Komple Excel'e Aktar", 
            font=ctk.CTkFont(size=14, weight="bold"),
            fg_color="#107C41",
            hover_color="#0A5A2F",
            height=45,
            command=self.sirketi_excele_donustur
        )
        btn_excel_aktar.pack(anchor="w", pady=(0, 5))

        # Bilgi Notu
        info_frame = ctk.CTkFrame(settings_frame, fg_color="transparent")
        info_frame.pack(pady=30, padx=30, fill="x")
        ctk.CTkLabel(info_frame, text="Ön Muhasebe Otomasyonu 1.6", text_color="gray").pack(side="right")

    def change_theme_event(self, new_theme):
        self.master_db.ayar_kaydet("tema", new_theme)
        ctk.set_appearance_mode(new_theme)
        if hasattr(self.main_app, 'apply_table_style'):
            self.main_app.apply_table_style()
        messagebox.showinfo("Tema Değişti", f"Görünüm modu '{new_theme}' olarak güncellendi.")

    def sirketi_excele_donustur(self):
        if not openpyxl:
            messagebox.showerror("Hata", "openpyxl kütüphanesi yüklü değil!\n'pip install openpyxl' komutu ile yükleyin.")
            return

        db = self.main_app.db
        db_path = db.db_name
        if not os.path.exists(db_path):
            messagebox.showerror("Hata", "Bu şirketin veritabanı bulunamadı.")
            return

        sirket_adi = os.path.splitext(os.path.basename(db_path))[0]
        tarih_str   = datetime.now().strftime("%Y%m%d_%H%M")
        varsayilan_ad = f"{sirket_adi}_TamVeriAktarim_{tarih_str}.xlsx"

        hedef_dosya = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=varsayilan_ad,
            title="Tüm Veriler — Excel Olarak Kaydet",
            filetypes=(("Excel Dosyası", "*.xlsx"), ("Tüm Dosyalar", "*.*"))
        )
        if not hedef_dosya:
            return

        # ── Stil Tanımları ─────────────────────────────────────────────────────
        RENK_BASLIK   = "1F4E79"   # Koyu lacivert (sol menü rengiyle uyumlu)
        RENK_GRUP     = "2E75B6"   # Orta mavi (alt başlık)
        RENK_STOK     = "375623"   # Koyu yeşil (stok)
        RENK_KASA     = "833C00"   # Kahverengi (kasa)
        RENK_BANKA    = "1B4F72"   # Deniz mavisi (banka)
        RENK_EVRAK    = "512DA8"   # Mor (evrak)
        RENK_PERSONEL = "4A235A"   # Koyu mor (personel)
        RENK_TAKVIM   = "1A5276"   # Petrol (takvim)
        RENK_NOT      = "34495E"   # Antrasit (notlar)
        THIN = Side(style="thin", color="BFBFBF")
        ince_border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

        def header_fill(hex_str):
            return PatternFill(start_color=hex_str, end_color=hex_str, fill_type="solid")

        def stil_uygula(ws, kolonlar, renk):
            """Başlık satırını formatla, sütun genişliklerini otomatik ayarla."""
            hf = header_fill(renk)
            hfont = Font(bold=True, color="FFFFFF", size=11)
            ws.append(kolonlar)
            for i, cell in enumerate(ws[ws.max_row], 1):
                cell.font  = hfont
                cell.fill  = hf
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = ince_border
                # Başlık genişliğine göre minimum genişlik
                ws.column_dimensions[get_column_letter(i)].width = max(16, len(str(kolonlar[i-1])) + 4)
            ws.row_dimensions[ws.max_row].height = 22

        def veri_yaz(ws, satirlar):
            """Satırları yaz ve zebra boyama + border uygula."""
            ac = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            for idx, satir in enumerate(satirlar, 1):
                ws.append([str(v) if v is not None else "" for v in satir])
                row_obj = ws[ws.max_row]
                for cell in row_obj:
                    if idx % 2 == 0:
                        cell.fill = ac
                    cell.border = ince_border
                    cell.alignment = Alignment(vertical="center")

        # ── Tüm sekmeler: (sekme_adı, SQL sorgusu, sütun başlıkları, renk) ───
        # Not: Sütun başlıkları elle yazılarak Türkçe okunabilirlik sağlanıyor
        with db._connect() as conn:
            cur = conn.cursor()

            def q(sql):
                try:
                    cur.execute(sql)
                    return cur.fetchall()
                except Exception as err:
                    print(f"SQL Hata: {err} | SQL: {sql}")
                    return []

            seksiyonlar = [
                # ── 1. CARİ YÖNETİMİ ──────────────────────────────────────────
                {
                    "baslik": "Cariler",
                    "renk": RENK_GRUP,
                    "kolonlar": ["ID", "Cari Ünvan", "Vergi No", "Vergi Dairesi", "Adres", "Telefon",
                                 "Bakiye (₺)", "Cari Türü", "Özel İskonto (%)", "Fiyat Grubu", "Kayıt Tarihi"],
                    "veri": q("""SELECT id, unvan, IFNULL(vergi_no,''), IFNULL(vergi_dairesi,''),
                                        IFNULL(adres,''), IFNULL(telefon,''), ROUND(IFNULL(bakiye,0),2),
                                        IFNULL(cari_turu,''), IFNULL(ozel_iskonto,0), IFNULL(fiyat_grubu,''),
                                        IFNULL(eklenme_tarihi,'')
                                 FROM cariler WHERE is_deleted=0 ORDER BY unvan""")
                },
                {
                    "baslik": "Cari Notları",
                    "renk": RENK_NOT,
                    "kolonlar": ["Not ID", "Cari Adı", "Not İçeriği", "Son Güncelleme"],
                    "veri": q("""SELECT cn.id, c.unvan, cn.not_icerik, cn.guncelleme_tarihi
                                 FROM cari_notlar cn
                                 LEFT JOIN cariler c ON cn.cari_id = c.id
                                 WHERE cn.is_deleted=0 ORDER BY c.unvan""")
                },
                # ── 2. STOK / ENVANTER ────────────────────────────────────────
                {
                    "baslik": "Stoklar",
                    "renk": RENK_STOK,
                    "kolonlar": ["ID", "Barkod", "Ürün Adı", "Miktar", "Birim", "Kritik Seviye", "Birim Fiyat (₺)"],
                    "veri": q("""SELECT id, IFNULL(barkod,''), urun_adi, ROUND(IFNULL(miktar,0),2),
                                        IFNULL(birim,'Adet'), IFNULL(kritik_seviye,0), ROUND(IFNULL(fiyat,0),2)
                                 FROM stoklar WHERE is_deleted=0 ORDER BY urun_adi""")
                },
                # ── 3. KASA YÖNETİMİ ─────────────────────────────────────────
                {
                    "baslik": "Kasalar",
                    "renk": RENK_KASA,
                    "kolonlar": ["ID", "Kasa Adı", "Kasa Bakiyesi (₺)"],
                    "veri": q("SELECT id, hesap_adi, ROUND(IFNULL(bakiye,0),2) FROM kasalar WHERE is_deleted=0")
                },
                {
                    "baslik": "Kasa Hareketleri",
                    "renk": RENK_KASA,
                    "kolonlar": ["ID", "Kasa", "Cari", "Tarih", "Tutar (₺)", "Tür", "Açıklama", "Proje Kodu"],
                    "veri": q("""SELECT kh.id, k.hesap_adi, IFNULL(c.unvan,'-'), kh.tarih,
                                        ROUND(kh.tutar,2), kh.tur, IFNULL(kh.aciklama,''), IFNULL(kh.proje_kodu,'')
                                 FROM kasa_hareketleri kh
                                 LEFT JOIN kasalar k ON kh.kasa_id = k.id
                                 LEFT JOIN cariler c ON kh.cari_id = c.id
                                 WHERE kh.is_deleted=0 ORDER BY kh.tarih DESC""")
                },
                # ── 4. BANKA YÖNETİMİ ────────────────────────────────────────
                {
                    "baslik": "Bankalar",
                    "renk": RENK_BANKA,
                    "kolonlar": ["ID", "Hesap Adı", "IBAN", "Bakiye (₺)"],
                    "veri": q("SELECT id, hesap_adi, IFNULL(iban,''), ROUND(IFNULL(bakiye,0),2) FROM bankalar WHERE is_deleted=0")
                },
                {
                    "baslik": "Banka Hareketleri",
                    "renk": RENK_BANKA,
                    "kolonlar": ["ID", "Banka", "Cari", "Tarih", "Tutar (₺)", "Tür", "Açıklama", "Proje Kodu"],
                    "veri": q("""SELECT bh.id, b.hesap_adi, IFNULL(c.unvan,'-'), bh.tarih,
                                        ROUND(bh.tutar,2), bh.tur, IFNULL(bh.aciklama,''), IFNULL(bh.proje_kodu,'')
                                 FROM banka_hareketleri bh
                                 LEFT JOIN bankalar b ON bh.banka_id = b.id
                                 LEFT JOIN cariler c ON bh.cari_id = c.id
                                 WHERE bh.is_deleted=0 ORDER BY bh.tarih DESC""")
                },
                # ── 5. MERKEZİ EVRAK (FATURA/İRSALİYE/SİPARİŞ) ─────────────
                {
                    "baslik": "Merkezi Evraklar",
                    "renk": RENK_EVRAK,
                    "kolonlar": ["ID", "Evrak No", "Cari", "Tarih", "Belge Türü", "Yön",
                                 "Ödeme Türü", "Toplam Tutar (₺)", "KDV Tutarı (₺)",
                                 "Tevkifat Türü", "Proje Kodu", "Açıklama"],
                    "veri": q("""SELECT f.id, f.fatura_no, IFNULL(c.unvan,'-'), f.tarih,
                                        IFNULL(f.belge_turu,'Fatura'), IFNULL(f.tur,''),
                                        IFNULL(f.odeme_turu,'Nakit'), ROUND(IFNULL(f.toplam_tutar,0),2),
                                        ROUND(IFNULL((SELECT SUM(fd2.miktar*fd2.birim_fiyat*IFNULL(fd2.kdv_orani,0)/100)
                                               FROM fatura_detay fd2 WHERE fd2.fatura_id=f.id AND fd2.is_deleted=0),0),2),
                                        IFNULL(f.tevkifat_turu,''), IFNULL(f.proje_kodu,''), IFNULL(f.aciklama,'')
                                 FROM faturalar f
                                 LEFT JOIN cariler c ON f.cari_id = c.id
                                 WHERE f.is_deleted=0 ORDER BY f.tarih DESC""")
                },
                {
                    "baslik": "Evrak Kalemleri",
                    "renk": RENK_EVRAK,
                    "kolonlar": ["ID", "Evrak No", "Cari", "Ürün Adı", "Miktar", "Birim",
                                 "Birim Fiyat (₺)", "KDV (%)", "Tevkifat (%)", "Satır Toplamı (₺)"],
                    "veri": q("""SELECT fd.id, f.fatura_no, IFNULL(c.unvan,'-'), IFNULL(s.urun_adi,''),
                                        ROUND(fd.miktar,2), IFNULL(fd.birim,''), ROUND(fd.birim_fiyat,2),
                                        IFNULL(fd.kdv_orani,0), IFNULL(fd.tevkifat_orani,0),
                                        ROUND(fd.miktar * fd.birim_fiyat, 2)
                                 FROM fatura_detay fd
                                 LEFT JOIN faturalar f ON fd.fatura_id = f.id
                                 LEFT JOIN stoklar s ON fd.stok_id = s.id
                                 LEFT JOIN cariler c ON f.cari_id = c.id
                                 WHERE fd.is_deleted=0 ORDER BY f.tarih DESC""")
                },
                # ── 6. PERSONEL TAKİBİ ────────────────────────────────────────
                {
                    "baslik": "Personel",
                    "renk": RENK_PERSONEL,
                    "kolonlar": ["ID", "Ad Soyad", "Kategori", "İşe Giriş",
                                 "İşten Çıkış", "Aylık Maaş (₺)", "Telefon", "Bakiye (₺)"],
                    "veri": q("""SELECT id, ad_soyad, IFNULL(kategori,''), IFNULL(ise_giris,''),
                                        IFNULL(isten_cikis,''), ROUND(IFNULL(aylik_maas,0),2),
                                        IFNULL(telefon,''), ROUND(IFNULL(bakiye,0),2)
                                 FROM calisanlar WHERE is_deleted=0 ORDER BY ad_soyad""")
                },
                {
                    "baslik": "Personel Hareketleri",
                    "renk": RENK_PERSONEL,
                    "kolonlar": ["ID", "Personel Adı", "Tarih", "Tutar (₺)", "İşlem Türü", "Açıklama"],
                    "veri": q("""SELECT ph.id, p.ad_soyad, ph.tarih, ROUND(ph.tutar,2), ph.tur, IFNULL(ph.aciklama,'')
                                 FROM personel_hareketleri ph
                                 LEFT JOIN calisanlar p ON ph.personel_id = p.id
                                 WHERE ph.is_deleted=0 ORDER BY ph.tarih DESC""")
                },
                # ── 7. AKILLI TAKVİM ──────────────────────────────────────────
                {
                    "baslik": "Takvim - Hatırlatıcılar",
                    "renk": RENK_TAKVIM,
                    "kolonlar": ["ID", "Başlık", "Açıklama", "Tarih", "Saat", "Öncelik"],
                    "veri": q("""SELECT id, baslik, IFNULL(aciklama,''), tarih, IFNULL(saat,''), IFNULL(oncelik,'Normal')
                                 FROM takvim_etkinlikleri WHERE is_deleted=0 ORDER BY tarih ASC""")
                },
            ]

        # ── Özet Sayfası ──────────────────────────────────────────────────────
        try:
            wb = openpyxl.Workbook()
            ws_ozet = wb.active
            ws_ozet.title = "📋 Özet"

            oz_baslik = Font(bold=True, size=16, color="FFFFFF")
            oz_fill   = header_fill("1F4E79")
            ws_ozet.append([f"Şirket Tam Veri Aktarımı — {sirket_adi}"])
            ws_ozet.append([f"Oluşturma Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M')}"])
            ws_ozet.append([])
            ws_ozet.append(["Sekme Adı", "Kayıt Sayısı"])
            for cell in ws_ozet[4]:
                cell.font = oz_baslik
                cell.fill = oz_fill
                cell.alignment = Alignment(horizontal="center")
            ws_ozet.column_dimensions["A"].width = 35
            ws_ozet.column_dimensions["B"].width = 18

            # Sayfaları oluştur
            for sek in seksiyonlar:
                ws = wb.create_sheet(title=sek["baslik"][:31])
                stil_uygula(ws, sek["kolonlar"], sek["renk"])
                veri_yaz(ws, sek["veri"])
                ws.freeze_panes = "A2"

                # Özet sayfasına satır ekle
                ws_ozet.append([sek["baslik"], len(sek["veri"])])

            wb.save(hedef_dosya)
            messagebox.showinfo(
                "✅ Aktarım Tamamlandı",
                f"Tüm modüller başarıyla Excel'e aktarıldı!\n\n"
                f"📁 Dosya: {os.path.basename(hedef_dosya)}\n"
                f"📊 Toplam Sekme: {len(seksiyonlar) + 1}\n\n"
                f"Kaydedildi: {hedef_dosya}"
            )

        except Exception as e:
            messagebox.showerror("Hata", f"Excel'e dönüştürme başarısız:\n{str(e)}")
